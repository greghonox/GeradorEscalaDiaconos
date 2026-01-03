"""Módulo para geração de planilha Excel da escala de diáconos."""

from datetime import date
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.gerador_escala import DiaconoEscala


class GeradorPlanilha:
    """Classe para gerar planilha Excel da escala de diáconos."""

    # Mapeamento de dias da semana para abreviações
    MAPA_DIAS_ABREV = {
        0: "S",  # Segunda
        1: "T",  # Terça
        2: "Q",  # Quarta
        3: "Q",  # Quinta
        4: "S",  # Sexta
        5: "S",  # Sábado
        6: "D",  # Domingo
    }

    # Mapeamento de dias para nomes de programas
    MAPA_PROGRAMAS = {
        "domingo": "Culto Evangelístico",
        "quarta": "Culto de Oração",
        "sabado": "Culto Divino",
    }

    # Mapeamento de nomes de meses
    MESES = [
        "Janeiro",
        "Fevereiro",
        "Março",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]

    def __init__(
        self,
        escala: List[DiaconoEscala],
        ano: int,
        lista_diaconos: List[Tuple[str, str]],
        seed: Optional[int] = None,
    ):
        """
        Inicializa o gerador de planilha.

        Args:
            escala: Lista com a escala gerada
            ano: Ano da escala
        """
        if not escala:
            raise ValueError("A escala não pode estar vazia")

        if not lista_diaconos:
            raise ValueError("A lista de diáconos não pode estar vazia")

        self.lista_diaconos_contatos = {
            nome: contato for nome, contato in lista_diaconos
        }

        self.escala = escala
        self.ano = ano
        self.wb = Workbook()
        # Remove a sheet padrão
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        # Gera mapeamento de cores para cada diácono
        self._cores_diaconos = self._gerar_cores_diaconos()
        self.seed = seed

    def _organizar_escala_por_data(
        self,
    ) -> Dict[date, Dict[str, Any]]:
        """
        Organiza a escala por data, agrupando chaves e ofertas.

        Returns:
            Dicionário com data como chave e dict com 'chave',
            'oferta' e 'dia' como listas/string
        """
        escala_por_data: Dict[date, Dict[str, Any]] = {}

        for diacono in self.escala:
            if diacono.data is None:
                continue

            if diacono.data not in escala_por_data:
                escala_por_data[diacono.data] = {
                    "chave": [],
                    "oferta": [],
                    "dia": diacono.dia,
                }

            escala_por_data[diacono.data][diacono.funcao].append(diacono.nome)

        return escala_por_data

    def _obter_meses_trimestre(self, trimestre: int) -> List[int]:
        """
        Retorna os meses de um trimestre (1-4).

        Args:
            trimestre: Número do trimestre (1, 2, 3 ou 4)

        Returns:
            Lista com os números dos meses (1-12)
        """
        if trimestre == 1:
            return [1, 2, 3]  # Janeiro a Março
        elif trimestre == 2:
            return [4, 5, 6]  # Abril a Junho
        elif trimestre == 3:
            return [7, 8, 9]  # Julho a Setembro
        elif trimestre == 4:
            return [10, 11, 12]  # Outubro a Dezembro
        else:
            raise ValueError("Trimestre deve ser 1, 2, 3 ou 4")

    def _filtrar_escala_por_trimestre(
        self,
        escala_por_data: Dict[date, Dict[str, Any]],
        trimestre: int,
    ) -> Dict[date, Dict[str, Any]]:
        """
        Filtra a escala para um trimestre específico.

        Args:
            escala_por_data: Escala organizada por data
            trimestre: Número do trimestre (1, 2, 3 ou 4)

        Returns:
            Escala filtrada para o trimestre
        """
        meses_trimestre = self._obter_meses_trimestre(trimestre)
        escala_filtrada = {}

        for data_evento, dados in escala_por_data.items():
            if data_evento.month in meses_trimestre:
                escala_filtrada[data_evento] = dados

        return escala_filtrada

    def _obter_nome_dia_abrev(self, weekday: int) -> str:
        """
        Retorna a abreviação do dia da semana.

        Args:
            weekday: Número do dia da semana (0=segunda, 6=domingo)

        Returns:
            Abreviação do dia
        """
        return self.MAPA_DIAS_ABREV.get(weekday, "")

    # Mapeamento de meses para colunas iniciais
    MAPA_COLUNAS_MESES = {
        "Janeiro": 1,  # A
        "Fevereiro": 6,  # G
        "Março": 11,  # M
        "Abril": 1,  # A
        "Maio": 6,  # G
        "Junho": 11,  # M
        "Julho": 1,  # A
        "Agosto": 6,  # G
        "Setembro": 11,  # M
        "Outubro": 1,  # A
        "Novembro": 6,  # G
        "Dezembro": 11,  # M
    }

    def _obter_coluna_mes(self, nome_mes: str) -> int:
        """
        Retorna o número da coluna inicial para o mês.

        Args:
            nome_mes: Nome do mês (ex: "Janeiro", "Fevereiro")

        Returns:
            Número da coluna (1=A, 7=G, 13=M, etc.)
        """
        return self.MAPA_COLUNAS_MESES.get(nome_mes, 1)

    def _gerar_cores_diaconos(self) -> Dict[str, str]:
        """
        Gera um mapeamento de cores únicas para cada diácono.

        Returns:
            Dicionário com nome do diácono como chave e cor hex como valor
        """
        # Lista de cores pastel e vibrantes para garantir boa visibilidade
        cores_disponiveis = [
            "FFB6C1",  # Light Pink
            "87CEEB",  # Sky Blue
            "98FB98",  # Pale Green
            "F0E68C",  # Khaki
            "DDA0DD",  # Plum
            "F5DEB3",  # Wheat
            "AFEEEE",  # Pale Turquoise
            "FFA07A",  # Light Salmon
            "20B2AA",  # Light Sea Green
            "FFD700",  # Gold
            "FF69B4",  # Hot Pink
            "00CED1",  # Dark Turquoise
            "FF6347",  # Tomato
            "7B68EE",  # Medium Slate Blue
            "32CD32",  # Lime Green
            "FF1493",  # Deep Pink
            "00BFFF",  # Deep Sky Blue
            "FFD700",  # Gold
            "ADFF2F",  # Green Yellow
        ]

        # Obtém todos os diáconos únicos da escala
        diaconos_unicos = set()
        for diacono in self.escala:
            diaconos_unicos.add(diacono.nome)

        diaconos_lista = sorted(list(diaconos_unicos))

        # Mapeia cada diácono a uma cor
        cores_diaconos: Dict[str, str] = {}
        for idx, nome_diacono in enumerate(diaconos_lista):
            # Usa cores disponíveis, repetindo se necessário
            cor = cores_disponiveis[idx % len(cores_disponiveis)]
            cores_diaconos[nome_diacono] = cor

        return cores_diaconos

    def _formatar_cabecalho_mes(self, ws, row: int, coluna_inicial: int):
        """
        Formata o cabeçalho de um mês específico.

        Args:
            ws: Worksheet do openpyxl
            row: Número da linha do cabeçalho
            coluna_inicial: Coluna inicial do mês
        """
        # Estilo do cabeçalho
        header_fill = PatternFill(
            start_color="808080", end_color="808080", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=8)
        header_alignment = Alignment(horizontal="center", vertical="center")
        background_fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )

        # Bordas
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        colunas = [
            "DATA",
            "DIA",
            "PROGRAMA",
            "RESPONSÁVEL (CHAVE)",
            "RESPONSÁVEL (OFERTA)",
        ]
        larguras = [5, 5, 15, 20, 20]
        altura = 20

        for offset, (coluna, largura) in enumerate(zip(colunas, larguras)):
            col_idx = coluna_inicial + offset
            cell = ws.cell(row=row, column=col_idx, value=coluna)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            cell.fill = background_fill
            ws.column_dimensions[get_column_letter(col_idx)].width = largura
            ws.row_dimensions[row].height = altura

    def _formatar_linha_dados(
        self,
        ws,
        row: int,
        coluna_inicial: int,
        chaves: Optional[List[str]] = None,
        ofertas: Optional[List[str]] = None,
        is_par: bool = False,
    ):
        """
        Formata uma linha de dados para um mês específico.

        Args:
            ws: Worksheet do openpyxl
            row: Número da linha
            coluna_inicial: Coluna inicial do mês
            chaves: Lista de nomes dos diáconos responsáveis pela chave
            ofertas: Lista de nomes dos diáconos responsáveis pela oferta
            is_par: Se True, aplica cor de fundo alternada
                (quando não há diácono específico)
        """
        chaves = chaves or []
        ofertas = ofertas or []

        # Cor de fundo padrão alternada (para células sem diácono específico)
        if is_par:
            fill_padrao = PatternFill(
                start_color="D3D3D3",
                end_color="D3D3D3",
                fill_type="solid",
            )
        else:
            fill_padrao = PatternFill(
                start_color="FFFFFF",
                end_color="FFFFFF",
                fill_type="solid",
            )

        data_height = 30
        # Bordas
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        data_font = Font(bold=True, color="000000", size=10)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        col_data_fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        col_data_font = Font(bold=True, color="FFFFFF", size=10)

        for offset in range(5):
            col = coluna_inicial + offset
            cell = ws.cell(row=row, column=col)

            # Coluna DATA (offset 0) sempre preta
            if offset == 0:
                cell.fill = col_data_fill
                cell.font = col_data_font
            # Coluna DIA (offset 1) - usa cor padrão
            elif offset == 1:
                cell.fill = fill_padrao
                cell.font = data_font
            # Coluna PROGRAMA (offset 2) - usa cor padrão
            elif offset == 2:
                cell.fill = fill_padrao
                cell.font = data_font
            # Coluna RESPONSÁVEL (CHAVE) (offset 3)
            elif offset == 3:
                if chaves:
                    # Se há múltiplos diáconos, usa a cor do primeiro
                    primeiro_diacono = chaves[0].strip()
                    if primeiro_diacono in self._cores_diaconos:
                        cor = self._cores_diaconos[primeiro_diacono]
                        cell.fill = PatternFill(
                            start_color=cor,
                            end_color=cor,
                            fill_type="solid",
                        )
                    else:
                        cell.fill = fill_padrao
                else:
                    cell.fill = fill_padrao
                cell.font = data_font
            # Coluna RESPONSÁVEL (OFERTA) (offset 4)
            elif offset == 4:
                if ofertas:
                    # Se há múltiplos diáconos, usa a cor do primeiro
                    primeiro_diacono = ofertas[0].strip()
                    if primeiro_diacono in self._cores_diaconos:
                        cor = self._cores_diaconos[primeiro_diacono]
                        cell.fill = PatternFill(
                            start_color=cor,
                            end_color=cor,
                            fill_type="solid",
                        )
                    else:
                        cell.fill = fill_padrao
                else:
                    cell.fill = fill_padrao
                cell.font = data_font

            cell.border = thin_border
            cell.alignment = alignment
            ws.row_dimensions[row].height = data_height

    def _adicionar_informacao_trimestre(self, ws):
        """
        Adiciona informação de horário de chegada do responsável.
        """
        row = 20
        information_text = "Horário de chegada do responsável deve ser pelo menos 30 minutos antes do horário marcado para o início de cada programação."
        ws.cell(row=row, column=1).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        ws.cell(row=row, column=1, value=information_text).font = Font(
            bold=True, size=12, color="FFFFFF"
        )
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        ws.merge_cells(f"A{row}:E24")

    def _adicionar_contato_responsavel(self, ws):
        """
        Adiciona informação de contato do responsável.
        Distribui os diáconos em duas colunas: Nome | Telefone, em blocos verticais e "quebra" para direita quando atinge o limite de linhas
        """
        column_start = 7
        max_rows = 20
        font_title = Font(bold=True, size=12, color="FFFFFF")
        background_title = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        font_name = Font(bold=True, size=10, color="000000")
        border_name = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        alignment_name = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=max_rows, column=column_start, value="Nome").font = font_title
        ws.cell(row=max_rows, column=column_start).border = border_name
        ws.cell(row=max_rows, column=column_start).alignment = alignment_name
        ws.cell(row=max_rows, column=column_start).fill = background_title

        ws.cell(row=max_rows, column=column_start + 2, value="Telefone").font = (
            font_title
        )
        ws.cell(row=max_rows, column=column_start + 2).border = border_name
        ws.cell(row=max_rows, column=column_start + 2).alignment = alignment_name
        ws.cell(row=max_rows, column=column_start + 2).fill = background_title
        ws.merge_cells(
            f"{get_column_letter(column_start)}{max_rows}:{get_column_letter(column_start + 1)}{max_rows}"
        )

        for idx, nome in enumerate(self.lista_diaconos_contatos.keys()):
            bloco = idx // max_rows
            pos = idx % max_rows
            row = 21 + pos
            column = column_start + 3 * bloco
            ws.cell(row=row, column=column, value=nome).font = font_name
            ws.cell(row=row, column=column).border = border_name
            ws.cell(row=row, column=column).alignment = alignment_name
            ws.cell(
                row=row, column=column + 2, value=self.lista_diaconos_contatos[nome]
            ).font = font_name
            ws.cell(row=row, column=column + 2).border = border_name
            ws.cell(row=row, column=column + 2).alignment = alignment_name
            ws.merge_cells(
                f"{get_column_letter(column)}{row}:{get_column_letter(column + 1)}{row}"
            )

    def _criar_sheet_trimestre(
        self,
        trimestre: int,
        escala_por_data: Dict[date, Dict[str, Any]],
    ):
        """
        Cria uma sheet para um trimestre específico.

        Args:
            trimestre: Número do trimestre (1, 2, 3 ou 4)
            escala_por_data: Escala organizada por data
        """
        meses_trimestre = self._obter_meses_trimestre(trimestre)
        nome_sheet = (
            f"{self.MESES[meses_trimestre[0] - 1]} a "
            f"{self.MESES[meses_trimestre[-1] - 1]}"
        )
        # Cria uma nova sheet para este trimestre
        ws = self.wb.create_sheet(title=nome_sheet)
        # Garante que a sheet foi criada corretamente
        if ws not in self.wb.worksheets:
            raise ValueError(f"Erro ao criar sheet: {nome_sheet}")

        # Filtra a escala para este trimestre
        escala_trimestre = self._filtrar_escala_por_trimestre(
            escala_por_data, trimestre
        )
        self._adicionar_informacao_trimestre(ws)
        self._adicionar_contato_responsavel(ws)

        # Título do trimestre - mescla todas as colunas dos meses
        primeira_coluna = self._obter_coluna_mes(self.MESES[meses_trimestre[0] - 1])
        ultima_coluna = self._obter_coluna_mes(self.MESES[meses_trimestre[-1] - 1])
        ultima_coluna_letra = get_column_letter(ultima_coluna + 4)
        primeira_coluna_letra = get_column_letter(primeira_coluna)
        ws.merge_cells(f"{primeira_coluna_letra}1:{ultima_coluna_letra}1")
        titulo_cell = ws.cell(
            row=1,
            column=primeira_coluna,
            value=f"{trimestre}° Trimestre/{self.ano}",
        )
        titulo_cell.font = Font(bold=True, size=16)
        titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
        mes_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Organiza por mês - cada mês começa em uma nova linha
        for mes in meses_trimestre:
            row = 2
            nome_mes = self.MESES[mes - 1]
            coluna_mes = self._obter_coluna_mes(nome_mes)
            coluna_letra = get_column_letter(coluna_mes)

            # Título do mês
            ultima_coluna_mes = get_column_letter(coluna_mes + 4)
            ws.merge_cells(f"{coluna_letra}{row}:{ultima_coluna_mes}{row}")
            mes_cell = ws.cell(row=row, column=coluna_mes, value=nome_mes)
            mes_cell.font = Font(bold=True, size=14)
            mes_cell.alignment = Alignment(horizontal="center", vertical="center")
            mes_cell.fill = PatternFill(
                start_color="E0E0E0",
                end_color="E0E0E0",
                fill_type="solid",
            )
            mes_cell.border = mes_border
            row += 1

            # Cabeçalho do mês
            self._formatar_cabecalho_mes(ws, row, coluna_mes)
            row += 1

            # Dados do mês
            datas_mes = [
                (data, dados)
                for data, dados in escala_trimestre.items()
                if data.month == mes
            ]
            datas_mes.sort()

            for data_evento, dados in datas_mes:
                # DATA
                ws.cell(row=row, column=coluna_mes, value=data_evento.day)

                # DIA
                dia_abrev = self._obter_nome_dia_abrev(data_evento.weekday())
                ws.cell(row=row, column=coluna_mes + 1, value=dia_abrev)

                # PROGRAMA
                dia_evento = dados.get("dia", "")
                programa = self.MAPA_PROGRAMAS.get(dia_evento, "")
                ws.cell(row=row, column=coluna_mes + 2, value=programa)

                # RESPONSÁVEL (CHAVE)
                chaves = dados.get("chave", [])
                if isinstance(chaves, list):
                    chave_str = " / ".join(chaves) if chaves else ""
                else:
                    chave_str = ""
                    chaves = []

                ws.cell(row=row, column=coluna_mes + 3, value=chave_str)

                # RESPONSÁVEL (OFERTA)
                ofertas = dados.get("oferta", [])
                if isinstance(ofertas, list):
                    oferta_str = " / ".join(ofertas) if ofertas else ""
                else:
                    oferta_str = ""
                    ofertas = []
                ws.cell(row=row, column=coluna_mes + 4, value=oferta_str)

                # Formata a linha com informações dos diáconos
                # para aplicar cores
                self._formatar_linha_dados(
                    ws,
                    row,
                    coluna_mes,
                    chaves=chaves if isinstance(chaves, list) else [],
                    ofertas=ofertas if isinstance(ofertas, list) else [],
                    is_par=(row % 2 == 0),
                )
                row += 1

            # Adiciona uma linha em branco entre meses para melhor visualização
            row += 1
        ws.cell(row=17, column=1, value=f"Seed: {self.seed}")

    def gerar_planilha(self, caminho_arquivo: str) -> None:
        """
        Gera a planilha Excel com 4 sheets (trimestres).

        Args:
            caminho_arquivo: Caminho onde salvar o arquivo Excel
        """
        escala_por_data = self._organizar_escala_por_data()

        # Cria os 4 sheets (trimestres)
        for trimestre in [1, 2, 3, 4]:
            self._criar_sheet_trimestre(trimestre, escala_por_data)

        # Salva o arquivo
        self.wb.save(caminho_arquivo)
